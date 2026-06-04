import tkinter as tk
from tkinter import ttk, font
import pandas as pd
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import matplotlib.gridspec as gridspec
from openpyxl import load_workbook

# ── Palette ────────────────────────────────────────────────────────────────
BG        = "#0F1117"
CARD_BG   = "#1A1D27"
CARD2_BG  = "#21253A"
ACCENT    = "#6C63FF"
ACCENT2   = "#00D4AA"
TEXT_PRI  = "#EAEAEA"
TEXT_SEC  = "#8A8FAA"
BORDER    = "#2C3050"

BLUES  = ["#3266AD","#4E86C8","#6AA0DE","#85B7EB","#A5CAF2","#C2DCF7","#DAEEFB"]
STATUS = ["#E24B4A","#F0997B","#EF9F27","#5DCAA5","#1D9E75"]
TEAL   = "#0F6E56"
PURPLE = "#7F77DD"

plt.rcParams.update({
    "figure.facecolor":  BG,
    "axes.facecolor":    BG,
    "axes.edgecolor":    BORDER,
    "axes.labelcolor":   TEXT_SEC,
    "xtick.color":       TEXT_SEC,
    "ytick.color":       TEXT_SEC,
    "text.color":        TEXT_PRI,
    "grid.color":        BORDER,
    "grid.linewidth":    0.5,
    "font.family":       "DejaVu Sans",
    "font.size":         10,
})

# ── Load data ───────────────────────────────────────────────────────────────
def load_data():
    wb = load_workbook(r"C:\Users\K\Downloads\Dashboard\Cleaned_Dataset.xlsx", read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    df = pd.DataFrame(rows[1:], columns=rows[0])
    df["TotalPrice"] = pd.to_numeric(df["TotalPrice"], errors="coerce")
    df["Quantity"]   = pd.to_numeric(df["Quantity"],   errors="coerce")
    return df

df = load_data()

status_counts   = df["OrderStatus"].value_counts()
product_rev     = df.groupby("Product")["TotalPrice"].sum().sort_values(ascending=False)
payment_counts  = df["PaymentMethod"].value_counts()
referral_counts = df["ReferralSource"].value_counts()
product_vol     = df["Product"].value_counts().sort_values()

# ── Root window ─────────────────────────────────────────────────────────────
root = tk.Tk()
root.title("E-Commerce Sales Dashboard")
root.configure(bg=BG)
root.geometry("1200x820")
root.minsize(1000, 700)

# ── Scrollable canvas ───────────────────────────────────────────────────────
outer = tk.Frame(root, bg=BG)
outer.pack(fill="both", expand=True)

canvas_scroll = tk.Canvas(outer, bg=BG, highlightthickness=0)
scrollbar = ttk.Scrollbar(outer, orient="vertical", command=canvas_scroll.yview)
canvas_scroll.configure(yscrollcommand=scrollbar.set)

scrollbar.pack(side="right", fill="y")
canvas_scroll.pack(side="left", fill="both", expand=True)

main = tk.Frame(canvas_scroll, bg=BG)
win_id = canvas_scroll.create_window((0, 0), window=main, anchor="nw")

def on_configure(e):
    canvas_scroll.configure(scrollregion=canvas_scroll.bbox("all"))
    canvas_scroll.itemconfig(win_id, width=canvas_scroll.winfo_width())

main.bind("<Configure>", on_configure)
canvas_scroll.bind("<Configure>", lambda e: canvas_scroll.itemconfig(win_id, width=e.width))
root.bind("<MouseWheel>", lambda e: canvas_scroll.yview_scroll(-1*(e.delta//120), "units"))
root.bind("<Button-4>",   lambda e: canvas_scroll.yview_scroll(-1, "units"))
root.bind("<Button-5>",   lambda e: canvas_scroll.yview_scroll( 1, "units"))

# ── Helpers ─────────────────────────────────────────────────────────────────
def section(parent, padx=16, pady=8):
    f = tk.Frame(parent, bg=BG)
    f.pack(fill="x", padx=padx, pady=pady)
    return f

def card(parent, col=0, colspan=1, row=0, rowspan=1, pad=6):
    f = tk.Frame(parent, bg=CARD_BG, bd=0, highlightbackground=BORDER,
                 highlightthickness=1)
    f.grid(row=row, column=col, columnspan=colspan, rowspan=rowspan,
           padx=pad, pady=pad, sticky="nsew")
    return f

def label(parent, text, size=11, color=TEXT_PRI, weight="normal", anchor="w", pady=0):
    tk.Label(parent, text=text, bg=parent["bg"], fg=color,
             font=("DejaVu Sans", size, weight), anchor=anchor,
             pady=pady).pack(fill="x", padx=12, pady=(4,0))

def embed_fig(parent, fig, height=240):
    c = FigureCanvasTkAgg(fig, master=parent)
    c.draw()
    w = c.get_tk_widget()
    w.configure(bg=BG, height=height)
    w.pack(fill="both", expand=True, padx=4, pady=(0,8))
    return c

# ═══════════════════════════════════════════════════════════════════════
# HEADER
# ═══════════════════════════════════════════════════════════════════════
hdr = tk.Frame(main, bg=CARD_BG, highlightbackground=BORDER, highlightthickness=1)
hdr.pack(fill="x", padx=16, pady=(16,8))

tk.Label(hdr, text="  ◈  E-Commerce Sales Dashboard", bg=CARD_BG, fg=TEXT_PRI,
         font=("DejaVu Sans", 15, "bold"), anchor="w", pady=12).pack(side="left", padx=6)
tk.Label(hdr, text="1,200 orders  •  7 products  •  5 channels",
         bg=CARD_BG, fg=TEXT_SEC, font=("DejaVu Sans", 10), pady=12).pack(side="right", padx=16)

# ═══════════════════════════════════════════════════════════════════════
# KPI CARDS
# ═══════════════════════════════════════════════════════════════════════
kpi_frame = section(main, pady=(4,4))
kpi_frame.columnconfigure((0,1,2,3), weight=1)

kpis = [
    ("Total Orders",    "1,200",     "ti-shopping-cart"),
    ("Total Revenue",   "$1,264,761","ti-currency-dollar"),
    ("Avg Order Value", "$1,054",    "ti-chart-line"),
    ("Unique Products", "7",         "ti-package"),
]

for i, (lbl, val, _) in enumerate(kpis):
    c = card(kpi_frame, col=i, pad=5)
    tk.Label(c, text=lbl, bg=CARD_BG, fg=TEXT_SEC,
             font=("DejaVu Sans", 9), anchor="w").pack(fill="x", padx=14, pady=(10,2))
    tk.Label(c, text=val, bg=CARD_BG, fg=ACCENT,
             font=("DejaVu Sans", 18, "bold"), anchor="w").pack(fill="x", padx=14, pady=(0,10))

# ═══════════════════════════════════════════════════════════════════════
# ROW 1 — Donut  +  Pie
# ═══════════════════════════════════════════════════════════════════════
row1 = section(main, pady=(4,4))
row1.columnconfigure((0,1), weight=1)

# — Order Status Donut —
c1 = card(row1, col=0)
label(c1, "Order Status Breakdown", size=10, color=TEXT_SEC)

fig1, ax1 = plt.subplots(figsize=(4.6, 2.8))
fig1.patch.set_facecolor(BG)
ax1.set_facecolor(BG)

wedges, texts, autotexts = ax1.pie(
    status_counts.values,
    labels=status_counts.index,
    colors=STATUS,
    autopct="%1.0f%%",
    pctdistance=0.78,
    startangle=90,
    wedgeprops=dict(width=0.52, edgecolor=BG, linewidth=2)
)
for t in texts:
    t.set(color=TEXT_SEC, fontsize=8)
for at in autotexts:
    at.set(color=TEXT_PRI, fontsize=7.5, fontweight="bold")
ax1.set_title("", pad=0)
embed_fig(c1, fig1, 240)

# — Revenue by Product Pie —
c2 = card(row1, col=1)
label(c2, "Revenue by Product", size=10, color=TEXT_SEC)

fig2, ax2 = plt.subplots(figsize=(4.6, 2.8))
fig2.patch.set_facecolor(BG)
ax2.set_facecolor(BG)

wedges2, texts2, autotexts2 = ax2.pie(
    product_rev.values,
    labels=product_rev.index,
    colors=BLUES,
    autopct="%1.0f%%",
    pctdistance=0.78,
    startangle=140,
    wedgeprops=dict(edgecolor=BG, linewidth=2)
)
for t in texts2:
    t.set(color=TEXT_SEC, fontsize=8)
for at in autotexts2:
    at.set(color=TEXT_PRI, fontsize=7.5, fontweight="bold")
embed_fig(c2, fig2, 240)

# ═══════════════════════════════════════════════════════════════════════
# ROW 2 — Payment  +  Referral horizontal bars
# ═══════════════════════════════════════════════════════════════════════
row2 = section(main, pady=(4,4))
row2.columnconfigure((0,1), weight=1)

def horiz_bar(parent, title, data, color, height=250):
    c = card(parent, col=0 if title.startswith("Payment") else 1)
    label(c, title, size=10, color=TEXT_SEC)
    fig, ax = plt.subplots(figsize=(4.6, 3.0))
    fig.patch.set_facecolor(BG)
    ax.set_facecolor(BG)
    bars = ax.barh(data.index, data.values, color=color, height=0.55,
                   edgecolor="none")
    for bar in bars:
        w = bar.get_width()
        ax.text(w + 2, bar.get_y() + bar.get_height()/2,
                str(int(w)), va="center", ha="left",
                color=TEXT_SEC, fontsize=8)
    ax.set_xlim(0, data.values.max() * 1.18)
    ax.tick_params(axis="y", labelsize=9, colors=TEXT_SEC)
    ax.tick_params(axis="x", labelsize=8, colors=TEXT_SEC)
    ax.xaxis.grid(True, linestyle="--", alpha=0.3)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(False)
    fig.tight_layout(pad=1.0)
    embed_fig(c, fig, height)
    return c

horiz_bar(row2, "Payment Methods",  payment_counts,  PURPLE, 250)
horiz_bar(row2, "Referral Sources", referral_counts, TEAL,   250)

# ═══════════════════════════════════════════════════════════════════════
# ROW 3 — Product Volume bar (full width)
# ═══════════════════════════════════════════════════════════════════════
row3 = section(main, pady=(4,12))
row3.columnconfigure(0, weight=1)

c5 = card(row3, col=0, colspan=1)
c5.configure(bg=CARD_BG)
label(c5, "Product Order Volume", size=10, color=TEXT_SEC)

fig5, ax5 = plt.subplots(figsize=(10, 2.6))
fig5.patch.set_facecolor(BG)
ax5.set_facecolor(BG)

vols  = df["Product"].value_counts().sort_values(ascending=False)
bars5 = ax5.bar(vols.index, vols.values, color=BLUES[:len(vols)],
                width=0.55, edgecolor="none")

for bar in bars5:
    h = bar.get_height()
    ax5.text(bar.get_x() + bar.get_width()/2, h + 1, str(int(h)),
             ha="center", va="bottom", fontsize=8.5, color=TEXT_SEC)

ax5.set_ylim(0, vols.values.max() * 1.15)
ax5.tick_params(axis="x", labelsize=9.5, colors=TEXT_SEC)
ax5.tick_params(axis="y", labelsize=8,   colors=TEXT_SEC)
ax5.yaxis.grid(True, linestyle="--", alpha=0.3)
ax5.set_axisbelow(True)
for spine in ax5.spines.values():
    spine.set_visible(False)

fig5.tight_layout(pad=1.0)
embed_fig(c5, fig5, 220)

# ═══════════════════════════════════════════════════════════════════════
# FOOTER
# ═══════════════════════════════════════════════════════════════════════
tk.Label(main, text="Project 4 — Data Visualization  •  1,200 orders",
         bg=BG, fg=TEXT_SEC, font=("DejaVu Sans", 8)).pack(pady=(0,12))

root.mainloop()
